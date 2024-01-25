import openpyxl


UPDATED_PRICES = {
    "Garlic": 3.07,
    "Celery": 1.19,
    "Lemon": 1.27
}


if __name__ == "__main__":

    print("Loading workbook...")
    wb = openpyxl.load_workbook("produceSales.xlsx")
    sheet = wb["Sheet"]

    print("Searching through products...")
    for row in range(2, sheet.max_row + 1):
        if (product := sheet.cell(row=row, column=1).value) in UPDATED_PRICES.keys():
            print(f"Changing price of {product}...")
            sheet.cell(row=row, column=2).value = UPDATED_PRICES[product]

    print("Saving updated file...")
    wb.save("updatedProduceSales.xlsx")

    print("Done.")
