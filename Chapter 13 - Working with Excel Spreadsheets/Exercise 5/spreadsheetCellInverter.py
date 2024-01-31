import sys

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


if __name__ == "__main__":

    if len(sys.argv) != 2:
        print("Usage: python spreadsheetCellInverter.py <filename>.xlsx")
        sys.exit()

    filename = sys.argv[1]

    try:
        wb = openpyxl.load_workbook(filename)
    except InvalidFileException:
        print("The given file DNE.")
        sys.exit()

    sheet = wb.active

    rows = []

    for row in range(1, sheet.max_row + 1):
        rows.append([])
        for col in range(1, sheet.max_column + 1):
            rows[row-1].append(sheet.cell(row=row, column=col).value)

    wb_inverted = openpyxl.Workbook()
    sheet = wb_inverted["Sheet"]

    for col, lst in enumerate(rows):
        for row, value in enumerate(lst):
            sheet.cell(row=row+1, column=col+1).value = rows[col][row]

    wb_inverted.save(f"inverted_{filename}")
    print("File inverted successfully.")
