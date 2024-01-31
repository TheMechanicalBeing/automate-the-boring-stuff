import sys

import openpyxl


if __name__ == "__main__":

    if len(sys.argv) != 4:
        print("Usage: python <filename>.py <integer> <integer> <filename>.xlsx")

    starting_row, followed_up, filename = sys.argv[1:]

    try:
        starting_row = int(starting_row)
        followed_up = int(followed_up)
    except ValueError:
        print("Second and third arguments must be integers")
        sys.exit()

    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    rows = []

    for row in range(1, sheet.max_row + 1):
        rows.append([])
        for col in range(1, sheet.max_column + 1):
            rows[row-1].append(sheet.cell(row=row, column=col).value)

    wb_to_save = openpyxl.Workbook()
    sheet = wb_to_save["Sheet"]

    row = 1
    for i in range(1, len(rows)+1):
        if i == starting_row:
            row += followed_up
        for col, value in enumerate(rows[i-1]):
            sheet.cell(row=row, column=col+1).value = value
        row += 1

    wb_to_save.save(f"updated_{filename}")
