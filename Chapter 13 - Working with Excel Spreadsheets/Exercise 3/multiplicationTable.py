import sys

import openpyxl
from openpyxl.styles import Font


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("You should provide exactly one argument.")
        sys.exit()

    try:
        dimension = int(sys.argv[1])
        if dimension < 1:
            raise ValueError()
    except ValueError:
        print("Argument should be a positive integer.")
        sys.exit()

    wb = openpyxl.Workbook()

    sheetname = f"{dimension}x{dimension}"
    wb.create_sheet(sheetname)
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheet = wb[sheetname]

    for col in range(2, dimension + 2):
        top_cell = sheet.cell(row=1, column=col)
        top_cell.value = col - 1
        top_cell.font = Font(bold=True)

    for row in range(2, dimension + 2):
        left_cell = sheet.cell(row=row, column=1)
        left_cell.value = row - 1
        left_cell.font = Font(bold=True)

        for col in range(2, dimension + 2):
            sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=1).value * sheet.cell(row=1, column=col).value

    wb.save("multiplicationTable.xlsx")
