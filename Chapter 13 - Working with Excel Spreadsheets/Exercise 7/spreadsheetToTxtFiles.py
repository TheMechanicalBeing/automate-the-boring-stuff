import sys

import openpyxl


if __name__ == "__main__":

    if len(sys.argv) != 2:
        print("Usage: python spreadsheetToTxtFiles.py <filename>.xlsx")
        sys.exit()

    filename = sys.argv[1]
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    for i in range(1, sheet.max_column+1):

        with open(f"txtFile{i}.txt", "w") as f:

            for j in range(1, sheet.max_row+1):
                line = sheet.cell(row=j, column=i).value
                line_to_write = line if line else ""
                f.write(line_to_write)

    print("Done")
