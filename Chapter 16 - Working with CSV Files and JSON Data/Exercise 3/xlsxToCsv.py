import os


import openpyxl
import csv


if __name__ == "__main__":
    for filename in os.listdir("."):
        if not filename.endswith(".xlsx"):
            continue

        wb = openpyxl.load_workbook(filename)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            csv_filename = f"./{filename.split('.')[0]}_{sheet_name}.csv"
            print(f"Writing data to {csv_filename}...")
            with open(csv_filename, "w", newline="") as f:
                writer = csv.writer(f)

                for row_num in range(1, sheet.max_row):
                    row_data = []

                    for column_num in range(1, sheet.max_column):
                        row_data.append(sheet.cell(row=row_num, column=column_num).value)

                    writer.writerow(row_data)
                print(f"Wrote {len(row_data)} records")

    print("Done")
