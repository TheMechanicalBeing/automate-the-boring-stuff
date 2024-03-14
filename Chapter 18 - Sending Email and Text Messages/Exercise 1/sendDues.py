import sys
import smtplib

import openpyxl


if __name__ == "__main__":
    wb = openpyxl.load_workbook('duesRecords.xlsx')
    sheet = wb.get_sheet_by_name("Sheet1")
    last_column = sheet.max_column
    latest_month = sheet.cell(row=1, column=last_column).value

    unpaid_members = {}
    for r in range(2, sheet.max_row + 1):
        payment = sheet.cell(row=r, column=last_column).value
        if payment != "paid":
            name = sheet.cell(row=r, column=1).value
            email = sheet.cell(row=r, column=2).value
            unpaid_members[name] = email

    smtp_object = smtplib.SMTP('smtp.example.com', 587)
    smtp_object.ehlo()
    smtp_object.starttls()
    smtp_object.login('MY_EMAIL', sys.argv[1])

    for name, email in unpaid_members.items():
        body = f"Subject: {latest_month} dues unpaid.\nDear {name},\nRecords show that you have not paid dues for {latest_month}. Please make this payment as soon as possible. Thank you!"
        print(f"Sending email to {email}...")
        sendmail_status = smtp_object.sendmail("MY_EMAIL", email, body)

        if sendmail_status != {}:
            print(f"There was a problem sending email to {email}: {sendmail_status}")
